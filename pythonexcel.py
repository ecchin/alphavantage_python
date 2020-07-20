from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
from openpyxl.styles import NamedStyle
import requests
import alpha_vantage
from xml.dom import minidom
from tkinter import messagebox
#import pandas as pd
import xlrd as x
import os 
from xlrd import open_workbook
from shutil import copyfile
from pathlib import Path
import os
import sys

#######################download data from alphavantage and create latest file
if len(sys.argv) == 1:
    configf = "config_daily.xml" 
elif len(sys.argv) == 2:
    if sys.argv[1] == "config_daily.xml":
        configf = "config_daily.xml"
    elif sys.argv[1] == "config_intraday.xml":
        configf = "config_intraday.xml"
    else:
        sys.exit("invalid config file") 
else:
    sys.exit("invalid config file")    


mydoc = minidom.parse(configf)
token = mydoc.getElementsByTagName('token')
function = mydoc.getElementsByTagName('function')
outputsize = mydoc.getElementsByTagName('outputsize')
interval = mydoc.getElementsByTagName('interval')


if function[0].firstChild.data == 'TIME_SERIES_DAILY':
    datastring = 'Time Series (Daily)'
    intervalstr = ""
else:
    datastring = 'Time Series (' + str(interval[0].firstChild.data) + ')'
    intervalstr = interval[0].firstChild.data, 

symbols = []
outer = []
dict = {}
lstofdct = []

file = open("stocklist.txt", "r") 
for line in file: 
    symbols.append(line.rstrip())
file.close()


date_style = NamedStyle(name='datetime', number_format='MM/DD/YYYY HH:MM:SS AM/PM')
date_style_d = NamedStyle(name='datetime_d', number_format='MM/DD/YYYY')
API_URL = "https://www.alphavantage.co/query" 
#symbols = ["BAC","MS","JPM"]
wb = Workbook()
allsheetsdictnew = {}

#use alphavantage to download data and save in excel file
s = 1
tt = 1
for symbol in symbols:
        data = { 
        "function": function[0].firstChild.data, 
        "symbol": symbol,
        "outputsize": outputsize[0].firstChild.data,
        "interval" : intervalstr,       
        "datatype": "json", 
        "apikey": token[0].firstChild.data } 
        response = requests.get(API_URL, data) 
        data = response.json()
        a = (data[datastring])
        keys = (a.keys())
        inner = []
        innerdct = {}
        
        if s == 1:
              ws0 = wb.active
              ws0.title = symbol
              ws0 = wb[symbol]
        else:
              ws0 = wb.create_sheet(symbol)
              ws0 = wb[symbol]
        r = 1
        for key in keys:
              if function[0].firstChild.data == 'TIME_SERIES_INTRADAY':
                  keyd = symbol + str(datetime.datetime.strptime(key, '%Y-%m-%d %H:%M:%S'))
              else:
                  keyd = symbol + str(datetime.datetime.strptime(key, '%Y-%m-%d'))
              allsheetsdictnew[keyd] = r
              for y in range(1,8):
                  if y == 1:
                      ws0.cell(row = r, column = y).value = symbol
                      inner.append(symbol)
                  elif y == 2:
                      if function[0].firstChild.data == 'TIME_SERIES_INTRADAY':
                           date_time_obj = datetime.datetime.strptime(key, '%Y-%m-%d %H:%M:%S')
                           ws0.cell(row = r, column = y).style = date_style
                      else:
                           date_time_obj = datetime.datetime.strptime(key, '%Y-%m-%d')
                           ws0.cell(row = r, column = y).style = date_style_d
                         
                      ws0.cell(row = r, column = y).value = date_time_obj
                      inner.append(date_time_obj)
                      innerdct[date_time_obj] = s
                  elif y == 3:
                      ws0.cell(row = r, column = y).value = float((a[key]['5. volume']))
                      inner.append(float((a[key]['5. volume'])))
                  elif y == 4:
                      ws0.cell(row = r, column = y).value = float((a[key]['1. open']))
                      inner.append(float((a[key]['1. open'])))
                  elif y == 5:
                      ws0.cell(row = r, column = y).value = float((a[key]['2. high']))
                      inner.append(float((a[key]['2. high'])))
                  elif y == 6:
                      ws0.cell(row = r, column = y).value = float((a[key]['3. low']))
                      inner.append(float((a[key]['3. low'])))
                  elif y == 7:
                      ws0.cell(row = r, column = y).value = float((a[key]['4. close']))
                      inner.append(float((a[key]['4. close'])))
              outer.append(inner)
              r = r + 1                   
        s = s + 1
        lstofdct.append(innerdct)




#print(len(outer))

if function[0].firstChild.data == 'TIME_SERIES_INTRADAY':
    intvstr = str(intervalstr).replace("(","").replace(")","").replace(",","").replace("'","") + "_"
else:
    intvstr = ""

now = datetime.datetime.now()
filestr = function[0].firstChild.data + "_" + intvstr + now.strftime("%m/%d/%Y, %H:%M:%S").replace(":", "_").replace("/", "_").replace(",", "_") + ".xlsx"
wb.save(filename = function[0].firstChild.data + "_" + intvstr + now.strftime("%m/%d/%Y, %H:%M:%S").replace(":", "_").replace("/", "_").replace(",", "_") + ".xlsx")
#######################################################################################################



########create master file from latest file if it doesn't exist, if it exists read through new data file, write to new merged file, create dictionary of new values
allsheetsdict = {}

if function[0].firstChild.data == 'TIME_SERIES_INTRADAY':
    pt = ("MASTER_ " + str(intervalstr) + ".xlsx").replace("',)", "").replace(" ('", "") 
    my_file = Path(pt)                 
    mfile  = pt
else:
    my_file = Path("MASTER_DAILY.xlsx")
    mfile = "MASTER_DAILY.xlsx"


if not (my_file.is_file()):
    if function[0].firstChild.data == 'TIME_SERIES_INTRADAY':
        pt = ("MASTER_ " + str(intervalstr) + ".xlsx").replace("',)", "").replace(" ('", "") 
        copyfile(filestr, pt)
    else:
        copyfile(filestr, "MASTER_DAILY.xlsx")
else:
    rowtrack = []
    shtnum = 1
    wbm2 = Workbook()
    wbc = open_workbook(filestr)
    sheets = wbc.sheets
    for sh in wbc.sheet_names():
        workbook = load_workbook(filestr)
        worksheet = workbook[sh]
        r = 1
        c = 1
        if shtnum == 1:
            ws0 = wbm2.active
            ws0.title = sh
            ws0 = wbm2[sh]
        else:
            ws0 = wbm2.create_sheet(sh)
            ws0 = wbm2[sh]
        for row_cells in worksheet.iter_rows():
            c = 1
            key = str(worksheet.cell(row = r, column = 1).value) + str(worksheet.cell(row = r, column = 2).value)          
            allsheetsdict[key] = r
            for cell in row_cells:
                 if c == 2:
                    if function[0].firstChild.data == 'TIME_SERIES_INTRADAY':
                        ws0.cell(row = r, column = c).style = date_style
                    else:
                        ws0.cell(row = r, column = c).style = date_style_d                    
                 ws0.cell(row = r, column = c).value = worksheet.cell(row = r, column = c).value
                 c = c +1
            r = r + 1
        shtnum = shtnum + 1
        rowtrack.append(r);
    
    #read master file and write entries not existing into new merged file, write new file out
    flag = 0    
    shtnum = 1
    wbc = open_workbook(mfile)
    sheets = wbc.sheets
    for sh in wbc.sheet_names():
        workbook2 = load_workbook(mfile)
        worksheet2 = workbook2[sh]
        ws0 = wbm2[sh]
        r = 1
        c = 1
        rt = rowtrack[shtnum]
        for row_cells in worksheet2.iter_rows():
            key = str(worksheet2.cell(row = r, column = 1).value) + str(worksheet2.cell(row = r, column = 2).value)
            c = 1
            for cell in row_cells:
                if c == 2:
                    if function[0].firstChild.data == 'TIME_SERIES_INTRADAY':
                        ws0.cell(row = rt, column = c).style = date_style
                    else:
                        ws0.cell(row = rt, column = c).style = date_style_d        
                if key not in allsheetsdict:             
                    ws0.cell(row = rt, column = c).value = worksheet2.cell(row = r, column = c).value
                    flag = 1
                c = c + 1
            if flag == 1:
                flag = 0
                rt = rt + 1
            r = r + 1
    wbm2.save(filename = "Mastermerged.xlsx")   
    os.remove(mfile)
    os.rename('Mastermerged.xlsx', mfile)

################################################################################################
